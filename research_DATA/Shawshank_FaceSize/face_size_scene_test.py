import csv 
import numpy as np 
import openpyxl
import re


path = '/Users/sohee/Documents/'
wb = openpyxl.load_workbook(path + 'shawshank_scene_facesize.xlsx')
ws = wb.active

####################################

scene_start = []

for col_J in ws["J"]:
	scene_start.append(str(col_J.value)) # 그냥 column.value로 하면 'cell'이라는 이상한 형식으로 저장됨. 그래서 다루기 쉽게 string으로 바꿔준 것.

while 'None' in scene_start:
	scene_start.remove('None')

scene_start.remove('시작 프레임')

print(scene_start)	
print("\n" + "-"*30)


# 여기까지가 씬 첫부분에 해당하는 프레임을 scene_start 리스트에. 

andy_frame = [] # 숫자 제거 전 프레임 이름 전체를 담은 리스트.

for col in ws.iter_cols(min_col=2, max_col=3):
	col = list(col)
	print(col)

for i in col[:5841]:
	andy_frame.append(i.value)
# while 'None' in andy_frame:
# 	andy_frame.remove(None)
print(andy_frame)
print(len(andy_frame))



'''
while 'None' in andy_frame:
	andy_frame.remove('None')

andy_frame_value = []

for frame in andy_frame:
	# print(frame)
	andy_frame_value.append(frame[10:15])

andy_frame_value.remove('s_uni')


# print(andy_frame)
print(andy_frame_value)
print(len(andy_frame_value))
print("\n" + "-"*30)

# 여기는 andy_frame열(C열) 전체를 andy_frame_value 리스트로. None 제거하고 숫자 부분만 담음.
'''
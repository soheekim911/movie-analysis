import csv 
import numpy as np 
import openpyxl
import re
import pdb


path = '/Users/sohee/Documents/research_DATA/Shawshank_Intensity/'

lines = []
scene_start = []

with open('shawshank_scene_arousal.csv', encoding='utf-8') as f:
	reader = csv.reader(f)
	for row in reader:
		print(row)
		lines.append(row)
		# pdb.set_trace()
		scene_start.append(row[1])

# print(lines[0])

# for row in lines[1:]:
# 	scene_start.append(float(int(row[''])))
# print(scene_start)

andy_arousal = [[0.0, 0.0] for i in range(len(scene_start))]


for row_line in lines[1:]:
	try:
		frame = int(row_line[2][21:26])
		for i in range(len(scene_start)-1):
			if frame >= int(scene_start[i]) and frame < int(scene_start[i+1]):
				#skeet_sizes[i].append(int(row_line[3]))
				andy_arousal[i][0] += int(row_line[3])
				andy_arousal[i][1] += 1
	except ValueError:
		pass


# for start, arousal in zip(scene_start, andy_size):
# 	print('start scene: {}, arousal level: {}'.format(start, andy_arousal))

			
#ws = wb.active
#
#####################################
#
#scene_start = []
#
#for col_J in ws["J"]:
#	scene_start.append(str(col_J.value)) # 그냥 column.value로 하면
# 'cell'이라는 이상한 형식으로 저장됨. 그래서 다루기 쉽게 string으로 바꿔준 것.
#
#while 'None' in scene_start:
#	scene_start.remove('None')
#
#scene_start.remove('시작 프레임')
#
#print(scene_start)	
#print("\n" + "-"*30)
#
#
## 여기까지가 씬 첫부분에 해당하는 프레임을 scene_start 리스트에. 
#
#skeet_frame = [] # 숫자 제거 전 프레임 이름 전체를 담은 리스트.
#
#for frame in ws["C"]:
#	frame = str(frame.value)
#	skeet_frame.append(frame)
#
#while 'None' in skeet_frame:
#	skeet_frame.remove('None')
#
#skeet_frame_value = []
#
#for frame in skeet_frame:
#	# print(frame)
#	skeet_frame_value.append(frame[10:15]) #'skeet_frame숫자.jpg'에서 숫자 부분만 남김
#
#skeet_frame_value.remove('s_uni')
#
#
## print(skeet_frame)
#print(skeet_frame_value)
#print(len(skeet_frame_value))
#print("\n" + "="*30)
#
## 여기는 skeet_frame열(C열) 전체를 skeet_frame_value 리스트로. None 제거하고 숫자 부분만 담음.
#
#
#for i in skeet_frame_value:
#	for size in ws["D"]:
#		size = str(size.value)
## dic = {}
## for i,j in skeet_frame_value, ws.columns[3]:
## 	dic[i] = j
## print(dic)
#
#scene_lists = []
#for start in scene_start:
#	scene_lists.append([]) # 리스트 안에 빈 리스트를 씬 갯수만큼 생성
#
#print(scene_lists)
#print(len(scene_lists))
#
#
#for frame in skeet_frame_value:
#	frame = int(frame)
#	# start = int(start)
#	for i in range(len(scene_start[:-1])):
#		if frame >= int(scene_start[i]) and frame < int(scene_start[i+1]):
#			scene_lists[i].append(frame)
#			print()
#
#print(scene_lists)
#
## facesize_skeet = []
#
## for col_D in ws["D"]:
## 	facesize_skeet.append(int(str(col_D.value)))
#
##for i in scene_lists:
##	for j in range(len(i)):
#		
#
## print(scene_lists)
## with open("scene_frame_list_skeet.txt", "w") as f:
## 	f.write(str(scene_lists))
#
## for i in scene_lists:
	# print(i)

	# for start in scene_start:
	# 	start = int(start)
	# 	if frame >= start and frame <= start 
	# 		scene_lists[i].append(frame)


# 여기는 scene start 값 < 인물 detect frame < scene start + 1 로 비교해서 각각 리스트화.



# for frame in skeet_frame_value:
# 	for i in len(scene_start):
# 		if frame > scene_start[i] and frame < scene_start[i+1]:
# 			skeet_start.append(frame)
	

#print((" %d 시 %d 분에 만나자")%(4,15))

# 여기는 skeet_frame_value 리스트의 값을 scene_start 값과 비교해서 각 씬에 속한 프레임들을 씬별 리스트화.



'''
하려는 것: 
1_ 일단 전체 프레임에 맞춰서 각 인물별로 detected된 프레임을 같은 열에 맞춰 나열. 
2_ 그리고 씬별 시작 프레임에 표시를 해서 그 부분부터 그 다음 표시까지가 하나의 씬 안에서 얼굴이 detected된 
프레임들만 모여있도록 할 거야. 
3_ 그렇게 해서 씬 별로 점수 내야지.

아니면 사실 1,2번은 내가 보기 편하려고 하는 거지 점수 내는 거랑은 상관 없으니까 하지 말고,
바로 '인물_frame00001.jpg'있는 열에서 정규표현식으로 씬 시작 프레임이랑 일치하는 부분부터 다음 일치 부분까지 확인해서
그 옆 열(실제로 detect된 face size)에서 해당 부분까지 점수 더해서 프레임 수로 나누면 되잖아.
그렇게 한번 코드 짜서 각 인물에 대해 적용하면 될 것 같은데.
'''

# pattern = re.compile("[a-z]+")

# for i in 
# result = pattern.findall


